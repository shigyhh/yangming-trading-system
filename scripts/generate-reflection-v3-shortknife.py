from __future__ import annotations

import copy
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
BASE_FILE = ROOT / "content/reflections/阳明心学交易系统-36场景360念统一调整表_照回V2戳心版.xlsx"
SELECTED_FILE = ROOT / "outputs/360-thoughts/阳明心学交易系统-360念照回终版_更狠戳心精选.xlsx"
OUTPUT_DIR = ROOT / "outputs/reflection-v3-shortknife"
OUTPUT_FILE = OUTPUT_DIR / "阳明心学交易系统-36场景360念统一调整表_照回V3短刀版.xlsx"

V3_COLUMNS = ["照回_V3短刀版", "刺痛等级_V3", "升级理由", "风险备注"]

THOUGHT_REWRITES = {
    "scene_06_010": "你松的不是利润。\n是终于不用再害怕。",
    "scene_09_001": "你补的不是仓位。\n是在给错误续命。",
}


EXACT_REWRITES = {
    "你说在复盘仓位，其实在反复地，责怪自己。": "你不是在复盘仓位。\n是在反复责怪自己。",
    "你说在复盘卖点，其实在为'贪'，又一次定罪。": "你不是在复盘卖点。\n是在给那次贪心定罪。",
    "你把'怕回吐'，当成了'见好就收'。": "你怕的不是回吐。\n是利润不肯留在你手里。",
    "便宜的是价格，没变的是你看错的方向。": "便宜的是价格。\n没变的是你看错的方向。",
    "你在给一个已经错的决定，一次次加注。": "你补的不是仓。\n是在给错误继续加注。",
    "这一刻你已经不在交易，你在赌。": "这一刻不是交易。\n是把结果交给一把赌。",
    "你盯的是成本线，不是这家公司，还值不值。": "你盯的不是价值。\n是成本线肯不肯放你走。",
    "你用别人的上涨，否定了自己的持仓。": "你换的不是股票。\n是被别人的上涨审判了。",
    "你开始找证据，去证明那条，你想买的消息。": "你找的不是证据。\n是消息替你下单的理由。",
    "'不敢贪'听起来克制，其实是，你不习惯赚钱。": "你不是不敢贪。\n是还不习惯拥有利润。",
    "你已经提前开始后悔一个还没发生的未来。": "你怕的不是持有。\n是提前后悔那个还没发生的未来。",
    "你难受的不是没上车。\n是别人已经在车上，而你还站在原地。": "没上车不疼。\n疼的是别人都走了，你还在原地。",
    "你怕的不是它起飞。\n你怕的是，自己又成了那个看着别人起飞的人。": "你怕的不是它起飞。\n是自己又成了看客。",
    "'应该没问题'。\n你不是判断的，是赌的，现在在求安心。": "你不是在判断。\n是在给那一把赌求安心。",
}


SCENE_FALLBACKS = {
    "急涨追高": "你追的不是机会。\n是怕这口肉没你的份。",
    "踏空焦虑": "没上车不疼。\n疼的是热闹里没有你的名字。",
    "买少后悔": "少买不疼。\n疼的是你没把幻想押满。",
    "卖飞懊恼": "你后悔的不是卖出。\n是没卖在幻想里的最高点。",
    "卖晚懊恼": "跌回来不疼。\n疼的是你看着贪心把利润拖没。",
    "浮盈回吐恐惧": "你怕的不是回撤。\n是到手的利润又不认你。",
    "不愿止损": "你不止损不是有信心。\n是怕亏损坐实：你错了。",
    "扛单死撑": "你扛的不是单。\n是那个不肯服输的自己。",
    "补仓摊平": "你补的不是仓位。\n是在给错误续命。",
    "越跌越补": "越补越不像策略。\n越像逼市场承认你没错。",
    "连续亏损后翻本": "你不是想交易。\n是想把羞耻一把打回去。",
    "连续盈利后膨胀": "账户涨的不是本事。\n是你开始以为市场服了你。",
    "重仓冲动": "你压的不是仓位。\n是想一笔证明自己。",
    "梭哈冲动": "你梭的不是机会。\n是想一把改命。",
    "空仓焦虑": "空仓不疼。\n疼的是热闹里没有你的名字。",
    "临盘改计划": "计划没坏。\n坏的是你先动了心。",
    "止损后又涨": "涨回来不疼。\n疼的是市场像在证明你错了。",
    "止盈后继续涨": "卖掉不疼。\n疼的是没吃到幻想里那一段。",
    "开盘冲动": "开盘没催你。\n是你的心先冲了出去。",
    "尾盘冲动": "尾盘不是机会。\n是不想空手收场。",
    "收盘后后悔": "你不是在复盘。\n是在把自己重新审一遍。",
    "复盘逃避": "你不是没时间。\n是怕证据摊开后无处可躲。",
    "消息刺激": "消息没那么重。\n重的是你怕慢别人一步。",
    "别人喊单依赖": "你问的不是方向。\n是想把责任先交出去。",
    "热点追逐": "热点没叫你。\n是你怕站错队。",
    "抄底冲动": "你抄的不是底。\n是想证明自己比市场聪明。",
    "高位接盘": "你接的不是机会。\n是被热闹推上了桌。",
    "回本执念": "回本不是出口。\n是你给错误要的体面。",
    "解套执念": "解套不是自由。\n是不想带着疼离开。",
    "看对不敢买": "你不是谨慎。\n是怕一出手又证明自己不行。",
    "买后立刻后悔": "你后悔的不是买入。\n是下单后才发现自己承受不起。",
    "频繁看账户": "你看的不是账户。\n是账户肯不肯给你一点安慰。",
    "频繁换股": "你换的不是股票。\n是受不了等待里的自己。",
    "消息面交易": "你交易的不是消息。\n是消息替你编出的想象。",
    "恐惧持有": "你怕的不是回落。\n是利润真的落到你手里。",
    "害怕认错": "亏损不最疼。\n疼的是承认那一刻：我错了。",
}


BANNED_SOFT_WORDS = ("其实", "说明", "反映", "代表", "意味着", "你需要", "你应该", "这是因为", "情绪", "认知偏差")
STRONG_WORDS = ("错", "羞耻", "责任", "体面", "证明", "赌", "改命", "背锅", "局外人", "审判", "服输", "放你走", "承受不起", "无处可躲")


def headers(ws):
    return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]


def row_dict(ws, row_index):
    h = headers(ws)
    return {name: ws.cell(row_index, i + 1).value for i, name in enumerate(h)}


def normalize_text(text: str) -> str:
    text = str(text or "").strip()
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = text.replace("，\n", "。\n")
    return text


def split_one_line(text: str) -> str:
    if "\n" in text:
        return text
    for sep in ("。是", "。你", "，是", "，而", "，没", "，不"):
        if sep in text:
            left, right = text.split(sep, 1)
            joiner = sep[-1]
            left = left.rstrip("。 ，")
            right = joiner + right
            return f"{left}。\n{right.rstrip('。')}。"
    return text


def shortknife(base_text: str, scene_name: str) -> tuple[str, bool]:
    base_text = normalize_text(base_text)
    if base_text in EXACT_REWRITES:
        return EXACT_REWRITES[base_text], True

    text = split_one_line(base_text)
    for word in BANNED_SOFT_WORDS:
        text = text.replace(word, "")
    text = text.replace("你是在", "是在")
    text = text.replace("你怕的是，", "是")
    text = text.replace("你等的是，", "是")
    text = text.replace("你追的是，", "是")
    text = text.replace("那个又一次", "那个又")
    text = text.replace("又一次", "又")
    text = text.replace("已经开始", "开始")
    text = normalize_text(text)

    compact_len = len(text.replace("\n", ""))
    has_bad_soft = any(word in text for word in BANNED_SOFT_WORDS)
    too_long = compact_len > 36
    too_explained = "，" in text and "\n" not in text
    if too_long or has_bad_soft or too_explained:
        fallback = SCENE_FALLBACKS.get(scene_name)
        if fallback:
            return fallback, True
    return text, text != base_text


def grade_v3(v3_text: str, selected_grade) -> int:
    try:
        grade = int(selected_grade)
    except Exception:
        grade = 4
    if any(word in v3_text for word in STRONG_WORDS):
        return 5
    return max(4, min(5, grade))


def risk_note(v3_text: str) -> str:
    compact = v3_text.replace("\n", "")
    if len(compact) > 42:
        return "太长，建议人工再压短。"
    if any(word in v3_text for word in ("蠢", "活该", "没用")):
        return "可能过度人身攻击，需人工复核。"
    if any(word in v3_text for word in BANNED_SOFT_WORDS):
        return "仍带解释词，需人工复核。"
    if "建议" in v3_text or "应该" in v3_text:
        return "像建议或训话，需人工复核。"
    return ""


def copy_header_style(ws, source_col: int, target_col: int):
    source = ws.cell(1, source_col)
    target = ws.cell(1, target_col)
    if source.has_style:
        target._style = copy.copy(source._style)
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.alignment = copy.copy(source.alignment)


def append_v3_columns(ws, selected_by_id):
    h = headers(ws)
    required = {"念ID"}
    if not required.issubset(set(h)):
        return []

    start_col = ws.max_column + 1
    for offset, name in enumerate(V3_COLUMNS):
        col = start_col + offset
        ws.cell(1, col).value = name
        copy_header_style(ws, ws.max_column if col == start_col else col - 1, col)
        ws.cell(1, col).fill = PatternFill("solid", fgColor="6F5423")
        ws.cell(1, col).font = Font(color="FFF2CC", bold=True)
        ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    index = {name: i + 1 for i, name in enumerate(h)}
    updated = []
    for row in range(2, ws.max_row + 1):
        thought_id = ws.cell(row, index["念ID"]).value
        selected = selected_by_id.get(thought_id)
        if not selected:
            continue
        if thought_id in THOUGHT_REWRITES:
            v3_text, changed = THOUGHT_REWRITES[thought_id], True
        else:
            v3_text, changed = shortknife(selected["最终照回"], selected["场景名称"])
        grade = grade_v3(v3_text, selected["刺痛等级"])
        reason = f"以{selected['采用来源']}的更狠精选为底稿，"
        reason += "去掉解释词并压成两拍短刀。" if changed else "保留原刀口，作为V3短刀定稿。"
        note = risk_note(v3_text)
        values = [v3_text, grade, reason, note]
        for offset, value in enumerate(values):
            cell = ws.cell(row, start_col + offset)
            cell.value = value
            if offset in (0, 2, 3):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        updated.append(
            {
                "念ID": thought_id,
                "场景名称": selected["场景名称"],
                "采用来源": selected["采用来源"],
                "原等级": selected["刺痛等级"],
                "刺痛等级_V3": grade,
                "风险备注": note,
            }
        )

    widths = [34, 12, 44, 32]
    for offset, width in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(start_col + offset)].width = width
    return updated


def main():
    selected_wb = openpyxl.load_workbook(SELECTED_FILE, data_only=True)
    selected_ws = selected_wb["照回终版精选"]
    selected_headers = headers(selected_ws)
    selected_index = {name: i + 1 for i, name in enumerate(selected_headers)}
    selected_by_id = {}
    for row in range(2, selected_ws.max_row + 1):
        thought_id = selected_ws.cell(row, selected_index["念ID"]).value
        selected_by_id[thought_id] = {
            "场景名称": selected_ws.cell(row, selected_index["场景名称"]).value,
            "最终照回": selected_ws.cell(row, selected_index["最终照回"]).value,
            "采用来源": selected_ws.cell(row, selected_index["采用来源"]).value,
            "刺痛点": selected_ws.cell(row, selected_index["刺痛点"]).value,
            "刺痛等级": selected_ws.cell(row, selected_index["刺痛等级"]).value,
        }

    wb = openpyxl.load_workbook(BASE_FILE)
    all_updates = []
    for sheet_name in ("360念统一表", "可导入_照回替换"):
        if sheet_name in wb.sheetnames:
            all_updates.extend(append_v3_columns(wb[sheet_name], selected_by_id))

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)

    main_updates = all_updates[: len(selected_by_id)]
    source_counts = Counter(item["采用来源"] for item in main_updates)
    grade_counts = Counter(item["刺痛等级_V3"] for item in main_updates)
    risk_rows = [item for item in main_updates if item["风险备注"]]
    report = {
        "output": str(OUTPUT_FILE),
        "processed": len(main_updates),
        "source_counts": dict(source_counts),
        "grade_counts": dict(grade_counts),
        "risk_count": len(risk_rows),
        "risk_rows": risk_rows[:30],
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
