from __future__ import annotations

import json
import random
import re
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs/reflection-v3-shortknife/阳明心学交易系统-36场景360念统一调整表_照回V3短刀版.xlsx"
REPORT = ROOT / "outputs/reflection-v3-shortknife/V3有效升级检查报告.md"

BANNED_TERMS = ["应该", "需要", "建议", "说明", "代表", "意味着", "因为", "其实", "情绪", "认知偏差"]
V3_COLUMNS = ["照回_V3短刀版", "刺痛等级_V3", "升级理由", "风险备注"]

THIEF_DIRECTIONS = {
    "贪": ["想拿全部", "嫌自己赚得不够", "把欲望伪装成机会"],
    "急": ["受不了等待", "把冲动叫果断", "手比规则快"],
    "痴": ["不肯认错", "把希望当判断", "让市场替自己洗错"],
    "执": ["放不下体面", "放不下成本", "让错误看起来没发生"],
    "疑": ["不敢承担", "还想再确认", "想要不会错的保证"],
    "怯": ["怕亏损坐实", "怕一出手就被打脸", "输不起"],
    "从": ["把责任交出去", "让别人替自己背结果", "躲进热闹里"],
    "傲": ["把运气当本事", "以为市场服了自己", "赢一次就飘"],
    "惧": ["怕到手的消失", "怕利润不属于自己", "用逃跑换安心"],
    "悔": ["拿过去审判自己", "把没吃到的当亏", "用后见之明折磨自己"],
    "懒": ["怕看见证据", "把复盘推给明天", "逃开失控现场"],
    "逃": ["躲开证据", "不想面对自己", "让问题先沉下去"],
    "赌": ["想一把改命", "把羞耻打回去", "把结果交给一把赌"],
}

SCENE_REWRITES = {
    "急涨追高": ["你追的不是机会。", "是怕这口肉没你的份。"],
    "踏空焦虑": ["没上车不疼。", "疼的是热闹里没有你的名字。"],
    "买少后悔": ["少买不疼。", "疼的是你没把幻想押满。"],
    "卖飞懊恼": ["卖掉不疼。", "疼的是没卖在幻想中的最高点。"],
    "卖晚懊恼": ["跌回来不疼。", "疼的是贪心把利润拖没。"],
    "浮盈回吐恐惧": ["你怕的不是回撤。", "是利润不肯留在你手里。"],
    "不愿止损": ["止损不疼。", "疼的是亏损坐实：你错了。"],
    "扛单死撑": ["你扛的不是仓。", "是那个不肯服输的自己。"],
    "补仓摊平": ["你补的不是仓位。", "是在给错误续命。"],
    "越跌越补": ["越补越不像策略。", "越像逼市场承认你没错。"],
    "连续亏损后翻本": ["你不是想交易。", "是想把羞耻一把打回去。"],
    "连续盈利后膨胀": ["账户涨的不是本事。", "是你开始以为市场服了你。"],
    "重仓冲动": ["你压的不是仓位。", "是想一笔证明自己。"],
    "梭哈冲动": ["你梭的不是机会。", "是想一把改命。"],
    "空仓焦虑": ["空仓不疼。", "疼的是热闹里没有你的名字。"],
    "临盘改计划": ["计划没坏。", "坏的是你先动了心。"],
    "止损后又涨": ["涨回来不疼。", "疼的是市场像在证明你错了。"],
    "止盈后继续涨": ["卖掉不疼。", "疼的是没吃到幻想里那一段。"],
    "开盘冲动": ["开盘没催你。", "是你的心先冲了出去。"],
    "尾盘冲动": ["尾盘不是机会。", "是不想空手收场。"],
    "收盘后后悔": ["你不是在复盘。", "是在把自己重新审一遍。"],
    "复盘逃避": ["你不是没时间。", "是怕证据摊开后无处可躲。"],
    "消息刺激": ["消息没那么重。", "重的是你怕慢别人一步。"],
    "别人喊单依赖": ["你问的不是方向。", "是想把责任先交出去。"],
    "热点追逐": ["热点没叫你。", "是你怕站错队。"],
    "抄底冲动": ["你抄的不是底。", "是想证明自己比市场聪明。"],
    "高位接盘": ["你接的不是机会。", "是被热闹推上了桌。"],
    "回本执念": ["回本不是出口。", "是你给错误要的体面。"],
    "解套执念": ["解套不是自由。", "是不想带着疼离开。"],
    "看对不敢买": ["你不是谨慎。", "是怕一出手又证明自己不行。"],
    "买后立刻后悔": ["买入不疼。", "疼的是下单后才发现承受不起。"],
    "频繁看账户": ["你看的不是账户。", "是账户肯不肯给你一点安慰。"],
    "频繁换股": ["你换的不是股票。", "是受不了等待里的自己。"],
    "消息面交易": ["你交易的不是消息。", "是消息替你编出的想象。"],
    "恐惧持有": ["你怕的不是回落。", "是利润真的落到你手里。"],
    "害怕认错": ["亏损不最疼。", "疼的是承认那一刻：我错了。"],
}

ROW_REWRITES = {
    "scene_03_008": "仓位不是问题。\n你在反复责怪那个没敢多买的自己。",
    "scene_11_004": "你要的不是机会。\n是市场还你一口气。",
    "scene_11_005": "再试一次不轻。\n里面藏着不甘心。",
    "scene_11_006": "差一点最会骗人。\n它让你继续把不甘押上去。",
    "scene_11_007": "你要弄回来的。\n是被亏损打碎的脸面。",
    "scene_11_008": "你不信一直亏。\n是不信自己会一直错。",
    "scene_11_010": "你没输给行情。\n你输不起那个结果。",
    "scene_21_002": "你折磨的不是当时。\n是那个停住的自己。",
    "scene_21_003": "你审的不是卖点。\n是那个先怕了的自己。",
    "scene_21_004": "本来能赚不疼。\n疼的是它只活在收盘后。",
    "scene_21_005": "你骂的不是这笔。\n是下单前那个失控的自己。",
    "scene_21_006": "白看的不是一天。\n是你把自己又晾在场外。",
    "scene_21_007": "亏赚都过去了。\n你还在用一天判自己。",
    "scene_21_008": "早知道不疼。\n疼的是你现在才敢承认。",
    "scene_21_009": "下次一定很轻。\n轻到遮不住这次没做到。",
    "scene_21_010": "真不该三个字。\n是在给自己补一刀。",
    "scene_22_001": "明天不是时间。\n是你给逃避找的门。",
    "scene_22_002": "累不是原因。\n是你不想再看见自己。",
    "scene_22_003": "没啥好记最危险。\n它把失控藏进空白。",
    "scene_22_004": "不看不是结束。\n是证据被你按进水底。",
    "scene_22_006": "过去的不是问题。\n是你不想让它留下证据。",
    "scene_22_007": "算了很轻。\n轻到刚好盖住失控。",
    "scene_22_008": "你没丢的是心情。\n你丢的是面对自己的勇气。",
    "scene_22_010": "明天开始。\n是今天继续逃开的借口。",
    "scene_30_003": "你要的不是确认。\n是有人替你承担结果。",
    "scene_30_004": "你怕的不是它起来。\n是自己又只看对、不敢做。",
    "scene_30_005": "该买不疼。\n疼的是你又没站到自己这边。",
    "scene_30_006": "他们敢不刺眼。\n刺眼的是你又退了半步。",
    "scene_30_008": "看飞的不是行情。\n是你那点不敢承担。",
    "scene_30_009": "计划没背叛你。\n是你先不敢认它。",
    "scene_30_010": "下次一定。\n是这次没敢买的遮羞布。",
    "scene_06_010": "你松的不是利润。\n是终于不用再害怕。",
    "scene_09_001": "你补的不是仓位。\n是在给错误续命。",
}


def normalize_exact(value):
    return str(value or "").strip().replace("\r\n", "\n").replace("\r", "\n")


def normalize_similarity(value):
    text = normalize_exact(value)
    return re.sub(r"[\s\n\r\t，。、“”‘’'\"：:；;！？!?（）()·,.]", "", text)


def compact_len(value):
    return len(re.sub(r"[\s\n\r\t]", "", str(value or "")))


def similarity(a, b):
    a_norm = normalize_similarity(a)
    b_norm = normalize_similarity(b)
    if not a_norm or not b_norm:
        return 0.0
    return SequenceMatcher(None, a_norm, b_norm).ratio()


def is_template(value):
    text = normalize_exact(value)
    lines = [line.strip() for line in text.split("\n") if line.strip()]
    if len(lines) >= 2 and lines[0].startswith("你不是") and re.match(r"^(是|是在|而是)", lines[1]):
        return True
    return bool(re.search(r"你不是[^。\n]{1,16}。\n(是|是在|而是)", text))


def split_thieves(value):
    text = str(value or "")
    return [part.strip() for part in re.split(r"[·、,/，\s]+", text) if part.strip()]


def headers(ws):
    return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]


def row_payload(ws, idx, row):
    return {
        "row": row,
        "念ID": ws.cell(row, idx["念ID"]).value,
        "场景名称": ws.cell(row, idx["场景名称"]).value,
        "心贼": ws.cell(row, idx["心贼"]).value,
        "表层念": ws.cell(row, idx["表层念"]).value,
        "隐藏念": ws.cell(row, idx["隐藏念"]).value,
        "原照回": ws.cell(row, idx["照回"]).value,
        "照回_v2": ws.cell(row, idx["照回_v2（戳心版）"]).value,
        "v3": ws.cell(row, idx["照回_V3短刀版"]).value,
    }


def metrics(ws, idx):
    exact_rows = []
    sim_rows = []
    long_rows = []
    banned_rows = []
    template_rows = []
    max_template_streak = 0
    current_streak = 0

    for row in range(2, ws.max_row + 1):
        payload = row_payload(ws, idx, row)
        v3 = payload["v3"]
        original = payload["原照回"]
        v2 = payload["照回_v2"]
        exact_hit = normalize_exact(v3) in {normalize_exact(original), normalize_exact(v2)}
        sim = max(similarity(v3, original), similarity(v3, v2))
        banned_hit = [term for term in BANNED_TERMS if term in str(v3 or "")]
        template_hit = is_template(v3)

        if exact_hit:
            exact_rows.append({**payload, "similarity": round(sim, 3)})
        if sim > 0.9:
            sim_rows.append({**payload, "similarity": round(sim, 3)})
        if compact_len(v3) > 28:
            long_rows.append({**payload, "length": compact_len(v3)})
        if banned_hit:
            banned_rows.append({**payload, "terms": banned_hit})
        if template_hit:
            template_rows.append(payload)
            current_streak += 1
            max_template_streak = max(max_template_streak, current_streak)
        else:
            current_streak = 0

    return {
        "total": ws.max_row - 1,
        "exact_count": len(exact_rows),
        "sim90_count": len(sim_rows),
        "long28_count": len(long_rows),
        "banned_count": len(banned_rows),
        "template_count": len(template_rows),
        "max_template_streak": max_template_streak,
        "exact_rows": exact_rows,
        "sim_rows": sim_rows,
        "long_rows": long_rows,
        "banned_rows": banned_rows,
        "template_rows": template_rows,
    }


def make_rewrite(payload):
    thought_id = payload["念ID"]
    if thought_id in ROW_REWRITES:
        return ROW_REWRITES[thought_id], "人工短刀覆写：修正别扭断句。"

    scene_name = payload["场景名称"]
    surface = str(payload["表层念"] or "").strip("。")
    hidden = str(payload["隐藏念"] or "").strip("。")
    thieves = split_thieves(payload["心贼"])
    first = thieves[0] if thieves else ""
    direction = THIEF_DIRECTIONS.get(first, ["照到背后的心贼"])[0]
    scene_lines = SCENE_REWRITES.get(scene_name)

    if scene_lines and len(surface) <= 7:
        # For very short OS, keep the scene knife. It tends to read cleaner than
        # mechanically repeating the user's words.
        return "\n".join(scene_lines), f"V3有效升级：从{first or '心贼'}入刀，压短到场景短刀。"

    if hidden and hidden not in surface:
        line1 = f"你嘴上说{surface}。"
        line2 = f"心里要的是{hidden}。"
    elif scene_lines:
        line1, line2 = scene_lines
    else:
        line1 = f"{surface}只是表面。"
        line2 = f"刀口在{direction}。"

    text = f"{line1}\n{line2}"
    for term in BANNED_TERMS:
        text = text.replace(term, "")
    text = text.replace("你嘴上说。", "你说不出口。")
    return text, f"V3有效升级：避开照搬，刀口落在{direction}。"


def apply_fixes(ws, idx, metric):
    problem_rows = {}
    for key in ("exact_rows", "sim_rows", "long_rows", "banned_rows", "template_rows"):
        for payload in metric[key]:
            problem_rows[payload["row"]] = payload

    for row, payload in problem_rows.items():
        new_text, reason = make_rewrite(payload)
        ws.cell(row, idx["照回_V3短刀版"]).value = new_text
        ws.cell(row, idx["照回_V3短刀版"]).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row, idx["升级理由"]).value = reason
        ws.cell(row, idx["风险备注"]).value = ""
        if compact_len(new_text) > 28:
            ws.cell(row, idx["风险备注"]).value = "超过28字，建议人工确认界面呈现。"

    return len(problem_rows)


def sample_by_thief(ws, idx):
    random.seed(20260609)
    grouped = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        payload = row_payload(ws, idx, row)
        payload["v3"] = ws.cell(row, idx["照回_V3短刀版"]).value
        payload["grade"] = ws.cell(row, idx["刺痛等级_V3"]).value
        for thief in split_thieves(payload["心贼"]):
            grouped[thief].append(payload)

    samples = {}
    for thief in sorted(grouped):
        items = grouped[thief]
        picked = random.sample(items, min(5, len(items)))
        samples[thief] = [
            {
                "念ID": item["念ID"],
                "场景": item["场景名称"],
                "表层念": item["表层念"],
                "V3": item["v3"],
                "等级": item["grade"],
            }
            for item in picked
        ]
    return samples


def write_report(before, after, fixed_count, samples):
    lines = [
        "# V3有效升级检查报告",
        "",
        "## 统计",
        "",
        "| 指标 | 修改前 | 修改后 |",
        "|---|---:|---:|",
        f"| 与原照回/照回_v2完全相同 | {before['exact_count']} | {after['exact_count']} |",
        f"| 与原照回/照回_v2相似度超过90% | {before['sim90_count']} | {after['sim90_count']} |",
        f"| 超过28个字 | {before['long28_count']} | {after['long28_count']} |",
        f"| 含解释/建议禁词 | {before['banned_count']} | {after['banned_count']} |",
        f"| 你不是X/是Y模板 | {before['template_count']} | {after['template_count']} |",
        f"| 模板最大连续段 | {before['max_template_streak']} | {after['max_template_streak']} |",
        f"| 本轮只改V3相关列条数 | {fixed_count} | - |",
        "",
        "## 每个心贼随机抽样",
        "",
    ]
    for thief, items in samples.items():
        lines.append(f"### {thief}")
        for item in items:
            v3 = str(item["V3"]).replace("\n", " / ")
            lines.append(f"- {item['念ID']}｜{item['场景']}｜{item['表层念']} -> {v3}（{item['等级']}）")
        lines.append("")
    REPORT.write_text("\n".join(lines), encoding="utf-8")


def sync_import_sheet(wb, source_ws, source_idx):
    if "可导入_照回替换" not in wb.sheetnames:
        return 0
    import_ws = wb["可导入_照回替换"]
    import_headers = headers(import_ws)
    import_idx = {name: i + 1 for i, name in enumerate(import_headers)}
    if "念ID" not in import_idx:
        return 0

    source_by_id = {}
    for row in range(2, source_ws.max_row + 1):
        thought_id = source_ws.cell(row, source_idx["念ID"]).value
        source_by_id[thought_id] = {
            column: source_ws.cell(row, source_idx[column]).value
            for column in V3_COLUMNS
        }

    synced = 0
    for row in range(2, import_ws.max_row + 1):
        thought_id = import_ws.cell(row, import_idx["念ID"]).value
        data = source_by_id.get(thought_id)
        if not data:
            continue
        for column in V3_COLUMNS:
            if column in import_idx:
                import_ws.cell(row, import_idx[column]).value = data[column]
                import_ws.cell(row, import_idx[column]).alignment = Alignment(vertical="top", wrap_text=True)
        synced += 1
    return synced


def main():
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["360念统一表"]
    h = headers(ws)
    idx = {name: i + 1 for i, name in enumerate(h)}
    missing = [name for name in ["照回", "照回_v2（戳心版）", *V3_COLUMNS] if name not in idx]
    if missing:
        raise RuntimeError(f"Missing columns: {missing}")

    before = metrics(ws, idx)
    fixed_count = apply_fixes(ws, idx, before)
    after = metrics(ws, idx)

    # If second pass still has problematic rows, leave them in risk notes rather
    # than touching non-V3 columns.
    for payload in after["long_rows"]:
        row = payload["row"]
        ws.cell(row, idx["风险备注"]).value = "超过28字，建议人工确认界面呈现。"
    for payload in after["banned_rows"]:
        row = payload["row"]
        terms = "、".join(payload["terms"])
        ws.cell(row, idx["风险备注"]).value = f"仍含解释/建议词：{terms}"

    samples = sample_by_thief(ws, idx)
    synced_import_rows = sync_import_sheet(wb, ws, idx)
    wb.save(WORKBOOK)
    write_report(before, after, fixed_count, samples)

    summary = {
        "workbook": str(WORKBOOK),
        "report": str(REPORT),
        "fixed_v3_rows": fixed_count,
        "synced_import_rows": synced_import_rows,
        "before": {k: before[k] for k in ["exact_count", "sim90_count", "long28_count", "banned_count", "template_count", "max_template_streak"]},
        "after": {k: after[k] for k in ["exact_count", "sim90_count", "long28_count", "banned_count", "template_count", "max_template_streak"]},
        "thieves": list(samples.keys()),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
