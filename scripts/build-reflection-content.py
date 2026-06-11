#!/usr/bin/env python3
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "content/reflections/reflection_final_神级择优导入版.xlsx"
SHEET_NAME = "可导入_神级择优版"
OUTPUT = ROOT / "web-next/src/content/reflections/reflection-final-shenji-zeyou.json"
VERSION = "reflection_final_shenji_zeyou_v1"


REQUIRED_COLUMNS = [
    "sceneId",
    "itemId",
    "tradeMoment",
    "os",
    "reflection_final",
    "final_source",
    "刺痛等级",
    "刺痛点",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def clean_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def main() -> None:
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")

    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_NAME}")

    sheet = workbook[SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet is empty: {SHEET_NAME}")

    headers = [clean_text(value) for value in rows[0]]
    column_index = {name: index for index, name in enumerate(headers) if name}
    missing = [name for name in REQUIRED_COLUMNS if name not in column_index]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    entries: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        payload = {name: row[index] if index < len(row) else None for name, index in column_index.items()}
        scene_id = clean_text(payload.get("sceneId"))
        item_id = clean_text(payload.get("itemId"))
        os_text = clean_text(payload.get("os"))
        reflection_final = clean_text(payload.get("reflection_final"))

        if not any([scene_id, item_id, os_text, reflection_final]):
            continue

        entry = {
            "sceneId": scene_id,
            "itemId": item_id,
            "tradeMoment": clean_text(payload.get("tradeMoment")),
            "os": os_text,
            "reflectionFinal": reflection_final,
            "finalSource": clean_text(payload.get("final_source")),
            "painLevel": clean_int(payload.get("刺痛等级")),
            "painPoint": clean_text(payload.get("刺痛点")),
            "sourceFile": clean_text(payload.get("sourceFile")),
            "version": VERSION,
        }

        final_reason = clean_text(payload.get("final_reason"))
        review_note = clean_text(payload.get("人工复核"))
        if final_reason:
            entry["finalReason"] = final_reason
        if review_note:
            entry["reviewNote"] = review_note

        if not scene_id or not item_id:
            raise ValueError(f"Row {row_number} is missing sceneId or itemId")

        entries.append(entry)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(entries, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Built {len(entries)} reflection entries -> {OUTPUT}")


if __name__ == "__main__":
    main()
